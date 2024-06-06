import re
import logging
from Lib.ParserUtils import ParserUtils
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)

class BatterySpecsInfo(MobileSpecInfo):
    def __init__(self):
        super().__init__()
        self.Battery_Level=None
    def grepInfo(self):
        self.getBatteryLevel()
        return self.BatterySpecsInfoDict
    def cleanup(self):
        pass

    def getBatteryLevel(self):
        ''' @function: getBatteryLevel gets Current Battery Level On Device
        @param: None
        @return: Current Battery Level on Device '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' dumpsys battery'
        self.Battery_Info = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r'level=(?P<level>\d+)')
        match = pattern.search(self.Battery_Info)
        if match:
            self.updateDictionary(dictName=self.HwSpecsInfoDict, key='Battery_Level', value=match.group('level'))
        return match.group('level')
    def generateXLSXReport(self, xlsObj=None, wb=None,ws=None, dataDict=None):
        headers = []
        headers.insert(0, "Parameters")
        headers.insert(1, "Results")

        for idx in range(0, len(headers)):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")

            cellref.value = headers[idx]


        dictkeys = list(dataDict.keys())
        for idx in range(0, len(dictkeys)):
            cellref = ws.cell(row=idx + 3, column=2)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            cellref.value = dictkeys[idx]

        #headers = list(dataDict.keys())
        col_idx = 3
        row_idx = 3
        for datavalue in dataDict.values():
            cellref = ws.cell(row=row_idx, column=col_idx)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            datavalue = FileSystemUtils.replaceChars(datavalue, charlist)
            cellref.value = str(datavalue)
            row_idx += 1

        col_idx = 2
        row_idx = len(list(dataDict.keys())) + 2
        for ctr in range(col_idx, col_idx + 2):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")

