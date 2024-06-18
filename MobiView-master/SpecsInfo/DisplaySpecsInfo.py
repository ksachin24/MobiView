import re
import logging
from Lib.ParserUtils import ParserUtils
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)


class DisplaySpecsInfo(MobileSpecsInfo):


    def grepInfo(self):

        self.getDisplayDensity()
        self.getScreenSize()
        self.getBrightness()
        self.getRefreshRate()
        self.getScreenOffTimeout()
        self.getScreenRotation()
        return self.DisplaySpecsInfoDict

    def cleanup(self):
        pass

    def getDisplayDensity(self):
        '''
        @function: getDevicePhysicalScreenDensity
            gets Device Physical Screen Density

        @param: None
        @return: Device Physical Screen Density
        '''
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' wm density '
        self.DislpayDensity = self.executeCommandOnDevice(command=self.command)
        # Need Parsing
        pattern = re.compile(r':\s+(?P<screen_density>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.DisplayDensity)

        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='DisplayDensity',
                              value=rvalue.get('screen_density'))
        return rvalue.get('screen_density')

    def getScreenSize(self):

        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' wm size '
        self.ScreenSize = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r':\s+(?P<screen_size>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.ScreenSize)
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='DisplayScreenSize',
                              value=rvalue.get('screen_size'))
        return rvalue.get('screen_size')

    def getBrightness(self):
        self.command = self.ADBObj.getADBDumpsysCommand() + ' settings get system screen_brightness '
        brightness = self.executeCommandOnDevice(command=self.command).strip()
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='Brightness', value=brightness)
        return brightness

    def getRefreshRate(self):
        self.command = self.ADBObj.getADBDumpsysCommand() + ' dumpsys display | grep "refreshRate" '
        refresh_rate = self.executeCommandOnDevice(command=self.command)
        for line in refresh_rate.splitlines():
            if "refreshRate" in line:
                rate = line.split("=")[1].strip()
                self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='Refresh_Rate', value=rate)
                return rate
        return None

    def getScreenOffTimeout(self):
        self.command = self.ADBObj.getADBDumpsysCommand() + ' settings get system screen_off_timeout '
        screen_off_timeout = self.executeCommandOnDevice(command=self.command).strip()
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='Screen_Off_Timeout', value=screen_off_timeout)
        return screen_off_timeout

    def getScreenRotation(self):
        self.command = self.ADBObj.getADBDumpsysCommand() + ' settings get system accelerometer_rotation '
        screen_rotation = self.executeCommandOnDevice(command=self.command).strip()
        self.updateDictionary(dictName=self.DisplaySpecsInfoDict, key='Screen_Rotation', value=screen_rotation)
        return screen_rotation

    def generateXLSXReport(self, xlsObj=None, wb=None, ws=None, dataDict=None):
        headers = ["Parameters", "Results"]

        for idx, header in enumerate(headers):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")
            cellref.value = header

        for idx, key in enumerate(dataDict.keys()):
            cellref = ws.cell(row=idx + 3, column=2)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            cellref.value = key

        for row_idx, value in enumerate(dataDict.values(), start=3):
            cellref = ws.cell(row=row_idx, column=3)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            value = FileSystemUtils.replaceChars(value, charlist)
            cellref.value = str(value)

        col_idx = 2
        row_idx = len(dataDict.keys()) + 3
        for ctr in range(col_idx, col_idx + 2):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")